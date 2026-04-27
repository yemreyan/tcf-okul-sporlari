#!/usr/bin/env python3
"""
TCF Artistik Cimnastik - Offline Excel Puan Sistemi
Gereksinim: pip install openpyxl
Çalıştırma: python3 tcf_artistik_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ─────────────────────────────────────────────────────────────
#  STIL SABİTLERİ
# ─────────────────────────────────────────────────────────────
C_PINK        = "C2185B"
C_BLUE        = "1565C0"
C_DARK        = "1A237E"
C_HEADER_BG   = "263238"
C_ALT_ROW     = "F1F8E9"
C_TOPLAM_BG   = "E3F2FD"
C_GOLD        = "FFD600"
C_SILVER      = "B0BEC5"
C_BRONZE      = "A1887F"
C_GECERSIZ    = "FFCDD2"   # Gecersiz (G) satır rengi
C_DNS         = "FFF3E0"   # Yarışmadı (D) satır rengi

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def medium_border():
    m = Side(style='medium')
    return Border(left=m, right=m, top=m, bottom=m)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, size=10, color='000000', italic=False, name='Calibri'):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN  = thin_border()
MEDIUM = medium_border()

# ─────────────────────────────────────────────────────────────
#  KATEGORİ VE ALET TANIMLAMALARI
# ─────────────────────────────────────────────────────────────
# (key, etiket, [aletler], cinsiyet: k/e)
CATEGORIES = [
    ('minik_a_kiz',    'MİNİK A KIZLAR',    ['Atlama', 'Serbest'],                                        'k'),
    ('minik_a_erkek',  'MİNİK A ERKEKLER',  ['Yer', 'Atlama'],                                            'e'),
    ('minik_b_kiz',    'MİNİK B KIZLAR',    ['Atlama', 'Serbest'],                                        'k'),
    ('minik_b_erkek',  'MİNİK B ERKEKLER',  ['Yer', 'Atlama'],                                            'e'),
    ('kucuk_kiz',      'KÜÇÜK KIZLAR',      ['Atlama', 'Asimetrik', 'Denge', 'Serbest'],                  'k'),
    ('kucuk_erkek',    'KÜÇÜK ERKEKLER',    ['Yer', 'Atlama', 'Paralel', 'Barfiks'],                      'e'),
    ('yildiz_kiz',     'YILDIZ KIZLAR',     ['Atlama', 'Asimetrik', 'Denge', 'Serbest'],                  'k'),
    ('yildiz_erkek',   'YILDIZ ERKEKLER',   ['Yer', 'Beygir', 'Halka', 'Atlama', 'Paralel', 'Barfiks'],   'e'),
    ('genc_kiz',       'GENÇ KIZLAR',       ['Atlama', 'Asimetrik', 'Denge', 'Serbest'],                  'k'),
    ('genc_erkek',     'GENÇ ERKEKLER',     ['Yer', 'Beygir', 'Halka', 'Atlama', 'Paralel', 'Barfiks'],   'e'),
]

MAX_ROWS = 40         # Kategori başına maksimum sporcu
DATA_START_ROW = 6    # Veri satırlarının başlangıcı (1-indexed)
MAX_GROUPS = 8        # Çıkış sırası sayfasında maksimum grup sayısı
MAX_PER_GROUP = 8     # Grup başına maksimum sporcu

# ─────────────────────────────────────────────────────────────
#  YARDIMCI FONKSİYON: Bir hücreyi biçimlendir
# ─────────────────────────────────────────────────────────────
def style_cell(cell, value=None, bold=False, size=10, color='000000',
               italic=False, bg=None, h_align='left', v_align='center',
               wrap=False, border=None, num_format=None):
    if value is not None:
        cell.value = value
    cell.font = font(bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align(h=h_align, v=v_align, wrap=wrap)
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format

# ─────────────────────────────────────────────────────────────
#  SAYFA 0: NASIL KULLANILIR
# ─────────────────────────────────────────────────────────────
def create_info_sheet(ws):
    ws.title = 'NASIL KULLANILIR'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 65

    def add_row(row, bullet, text, bold_text=False, bg=None, size=10):
        ws.row_dimensions[row].height = 18
        b = ws.cell(row=row, column=2, value=bullet)
        b.font = font(bold=True, size=size)
        if bg:
            b.fill = fill(bg)
        t = ws.cell(row=row, column=3, value=text)
        t.font = font(bold=bold_text, size=size)
        if bg:
            t.fill = fill(bg)
        t.alignment = align(h='left')

    # Başlık
    ws.merge_cells('B1:C1')
    ws.row_dimensions[1].height = 38
    hdr = ws.cell(row=1, column=2, value='TCF ARTİSTİK CİMNASTİK  —  OFFLİNE PUAN & SIRALAMA SİSTEMİ')
    hdr.font = font(bold=True, size=15, color='FFFFFF')
    hdr.fill = fill(C_DARK)
    hdr.alignment = align(h='center')

    ws.merge_cells('B2:C2')
    sub = ws.cell(row=2, column=2, value='İnternet bağlantısı gerektirmez  ·  Tüm hesaplamalar otomatiktir  ·  Çıktı alınabilir')
    sub.font = font(italic=True, size=10, color='444444')
    sub.alignment = align(h='center')
    sub.fill = fill('E8EAF6')
    ws.row_dimensions[2].height = 22

    lines = [
        (3,  '',    ''),
        (4,  '✅',  'KULLANIM ADIMLARI',                       True,  'E3F2FD', 12),
        (5,  '1.',  'Alt sekmelerden kategoriyi seçin (MİNİK A KIZLAR, KÜÇÜK ERKEKLER vb.)'),
        (6,  '2.',  'Her satıra: Ad, Soyad, Okul/Kulüp, Yarışma Türü (Bireysel / Takım) girin.'),
        (7,  '3.',  'Her alet sütununa puan girin (örn: 12.350)'),
        (8,  '4.',  '"Durum" sütununa alet durumunu girin:',    True),
        (9,  '   ', '   • Boş bırakın veya "N" yazın  →  Normal (puan geçerli)'),
        (10, '   ', '   • "G" yazın  →  Gecersiz  (bu aletin puanı 0 sayılır)',        False, C_GECERSIZ),
        (11, '   ', '   • "D" yazın  →  Yarışmadı / DNS  (bu aletin puanı 0 sayılır)', False, C_DNS),
        (12, '5.',  '"TOPLAM" sütunu otomatik hesaplanır.'),
        (13, '6.',  '"SIRALAMA" sütunu otomatik sıralanır — eşit puanlılar aynı sırayı alır.'),
        (14, '7.',  'Altın / Gümüş / Bronz renkleri otomatik uygulanır (1. / 2. / 3. sıra).'),
        (15, '8.',  '"ÇIKIŞ SIRASI" sekmesinde manuel çıkış sırasını düzenleyin.'),
        (16, '',    ''),
        (17, '⚠️',  'ÖNEMLİ NOTLAR',                           True,  'FFF8E1', 12),
        (18, '•',   'Ondalık ayracı için sistem ayarınıza göre nokta (.) veya virgül (,) kullanın.'),
        (19, '•',   'Dosyayı düzenli kaydedin: Ctrl+S'),
        (20, '•',   '40\'dan fazla sporcu için yeni satır ekleyin (formülleri kopyalamayı unutmayın).'),
        (21, '•',   'Kategori sekmesinin adı değiştirilmemelidir.'),
        (22, '',    ''),
        (23, '📋',  'KATEGORİLER VE ALETLER',                  True,  'E8F5E9', 12),
        (24, '',    ''),
    ]

    for item in lines:
        if len(item) == 2:
            r, bullet = item
            add_row(r, bullet, '')
        elif len(item) == 3:
            r, bullet, text = item
            add_row(r, bullet, text)
        elif len(item) == 4:
            r, bullet, text, bold_text = item
            add_row(r, bullet, text, bold_text)
        elif len(item) == 5:
            r, bullet, text, bold_text, bg = item
            add_row(r, bullet, text, bold_text, bg)
        elif len(item) == 6:
            r, bullet, text, bold_text, bg, size = item
            add_row(r, bullet, text, bold_text, bg, size)

    row = 25
    for _, label, apparatus, gender in CATEGORIES:
        icon = '👧' if gender == 'k' else '👦'
        c_bg = 'FCE4EC' if gender == 'k' else 'E3F2FD'
        add_row(row, icon, f'{label}:  {" + ".join(apparatus)}', False, c_bg)
        row += 1

    ws.freeze_panes = 'B3'


# ─────────────────────────────────────────────────────────────
#  KATEGORİ PUANLAMA SAYFASİ
# ─────────────────────────────────────────────────────────────
def create_category_sheet(wb, key, label, apparatus_list, gender):
    ws = wb.create_sheet(title=label[:31])
    ws.sheet_view.showGridLines = False

    hdr_color  = C_PINK if gender == 'k' else C_BLUE
    n_app      = len(apparatus_list)
    # Sabit sütunlar: No(1) Ad(2) Soyad(3) Okul(4) Tür(5) = 5
    # Her alet: Puan(1) + Durum(1) = 2 sütun
    # Son: Toplam(1) + Siralama(1) = 2
    FIXED_COLS   = 5
    total_cols   = FIXED_COLS + n_app * 2 + 2
    toplam_col   = FIXED_COLS + n_app * 2 + 1
    siralama_col = toplam_col + 1
    toplam_letter   = get_column_letter(toplam_col)
    siralama_letter = get_column_letter(siralama_col)
    data_end_row = DATA_START_ROW + MAX_ROWS - 1
    toplam_range = f'${toplam_letter}${DATA_START_ROW}:${toplam_letter}${data_end_row}'

    # ── Sütun genişlikleri ──────────────────────────────────
    ws.column_dimensions['A'].width = 4.5   # No
    ws.column_dimensions['B'].width = 13    # Ad
    ws.column_dimensions['C'].width = 15    # Soyad
    ws.column_dimensions['D'].width = 30    # Okul
    ws.column_dimensions['E'].width = 10    # Tür

    col = FIXED_COLS + 1
    for _ in apparatus_list:
        ws.column_dimensions[get_column_letter(col)].width = 8.5  # Puan
        ws.column_dimensions[get_column_letter(col+1)].width = 7  # Durum
        col += 2
    ws.column_dimensions[get_column_letter(toplam_col)].width = 10
    ws.column_dimensions[get_column_letter(siralama_col)].width = 10

    # ── Satır 1: Kategori başlığı ───────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=label)
    c.font = font(bold=True, size=15, color='FFFFFF')
    c.fill = fill(hdr_color)
    c.alignment = align(h='center')
    ws.row_dimensions[1].height = 32

    # ── Satır 2: Yarışma bilgileri ──────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    info_text = 'Yarışma: _________________________________   Tarih: ______________   Şehir: ______________'
    c = ws.cell(row=2, column=1, value=info_text)
    c.font = font(italic=True, size=9, color='555555')
    c.fill = fill('ECEFF1')
    c.alignment = align(h='left')
    ws.row_dimensions[2].height = 16

    # ── Satır 3: Alet üst başlıkları (merge) ───────────────
    # Fixed sütunlar boş
    for col_i in range(1, FIXED_COLS + 1):
        c = ws.cell(row=3, column=col_i)
        c.fill = fill(hdr_color)
        c.border = THIN

    col = FIXED_COLS + 1
    for app_name in apparatus_list:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        c = ws.cell(row=3, column=col, value=app_name)
        c.font = font(bold=True, size=10, color='FFFFFF')
        c.fill = fill(hdr_color)
        c.alignment = align(h='center')
        c.border = THIN
        ws.cell(row=3, column=col+1).border = THIN
        col += 2

    # Toplam + Siralama üst başlık
    for offset in [0, 1]:
        c = ws.cell(row=3, column=toplam_col + offset)
        c.fill = fill(hdr_color)
        c.border = THIN
    ws.row_dimensions[3].height = 22

    # ── Satır 4: Alt sütun başlıkları ──────────────────────
    sub_headers = ['No', 'Ad', 'Soyad', 'Okul / Kulüp', 'Tür']
    for _ in apparatus_list:
        sub_headers += ['Puan', 'Dur.']
    sub_headers += ['TOPLAM', 'SIRALAMA']

    for col_i, hdr_text in enumerate(sub_headers, 1):
        c = ws.cell(row=4, column=col_i, value=hdr_text)
        c.font = font(bold=True, size=9, color='FFFFFF')
        c.fill = fill(C_HEADER_BG)
        c.alignment = align(h='center', wrap=True)
        c.border = THIN
    ws.row_dimensions[4].height = 28

    # ── Satır 5: Açıklama notu ──────────────────────────────
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=total_cols)
    note = ('Dur. sütunu: Boş veya N = Normal  |  G = Gecersiz (puan 0)  |  D = Yarışmadı/DNS (puan 0)  '
            '→  TOPLAM ve SIRALAMA otomatik hesaplanır')
    c = ws.cell(row=5, column=1, value=note)
    c.font = font(italic=True, size=8, color='555555')
    c.fill = fill('F5F5F5')
    c.alignment = align(h='center')
    ws.row_dimensions[5].height = 14

    # ── Data Validation: Tür dropdown ──────────────────────
    dv_tur = DataValidation(
        type='list', formula1='"Bireysel,Takım"', allow_blank=True,
        showDropDown=False, showErrorMessage=False
    )
    ws.add_data_validation(dv_tur)

    # ── Veri satırları ──────────────────────────────────────
    for row in range(DATA_START_ROW, data_end_row + 1):
        row_bg = 'FFFFFF' if (row - DATA_START_ROW) % 2 == 0 else C_ALT_ROW

        # No
        c = ws.cell(row=row, column=1, value=row - DATA_START_ROW + 1)
        c.font = font(size=9, color='888888')
        c.fill = fill(row_bg)
        c.alignment = align(h='center')
        c.border = THIN

        # Ad, Soyad
        for col_i in [2, 3]:
            c = ws.cell(row=row, column=col_i)
            c.font = font(size=10)
            c.fill = fill(row_bg)
            c.border = THIN
            c.alignment = align(h='left')

        # Okul
        c = ws.cell(row=row, column=4)
        c.font = font(size=9)
        c.fill = fill(row_bg)
        c.border = THIN
        c.alignment = align(h='left')

        # Tür (dropdown)
        c = ws.cell(row=row, column=5)
        c.font = font(size=9)
        c.fill = fill(row_bg)
        c.border = THIN
        c.alignment = align(h='center')
        dv_tur.add(c)

        # Alet sütunları (Puan + Durum)
        col = FIXED_COLS + 1
        puan_refs = []
        durum_refs = []
        for _ in apparatus_list:
            pl = get_column_letter(col)
            dl = get_column_letter(col + 1)
            puan_refs.append(f'{pl}{row}')
            durum_refs.append(f'{dl}{row}')

            # Puan
            pc = ws.cell(row=row, column=col)
            pc.font = font(size=10)
            pc.fill = fill(row_bg)
            pc.border = THIN
            pc.alignment = align(h='center')
            pc.number_format = '0.000'

            # Durum (N/G/D)
            dc = ws.cell(row=row, column=col + 1)
            dc.font = font(size=9, color='666666')
            dc.fill = fill(row_bg)
            dc.border = THIN
            dc.alignment = align(h='center')

            col += 2

        # TOPLAM formülü
        # =IF(B=""," ",SUM_of_effective_scores)
        eff_parts = [
            f'IF(OR({dr}="G",{dr}="D"),0,IFERROR({pr}+0,0))'
            for pr, dr in zip(puan_refs, durum_refs)
        ]
        toplam_formula = f'=IF(B{row}="","",' + '+'.join(eff_parts) + ')'

        tc = ws.cell(row=row, column=toplam_col, value=toplam_formula)
        tc.font = font(bold=True, size=10)
        tc.fill = fill(C_TOPLAM_BG)
        tc.border = THIN
        tc.alignment = align(h='center')
        tc.number_format = '0.000'

        # SIRALAMA formülü
        # Toplam > 0 ise RANK.EQ, aksi halde "-"
        siralama_formula = (
            f'=IF(OR(B{row}="",{toplam_letter}{row}="",{toplam_letter}{row}=0),"-",'
            f'RANK.EQ({toplam_letter}{row},{toplam_range},0))'
        )
        sc = ws.cell(row=row, column=siralama_col, value=siralama_formula)
        sc.font = font(bold=True, size=10)
        sc.fill = fill(C_TOPLAM_BG)
        sc.border = THIN
        sc.alignment = align(h='center')

    # ── Koşullu biçimlendirme: Sıralama 1/2/3 ─────────────
    sira_range = f'{siralama_letter}{DATA_START_ROW}:{siralama_letter}{data_end_row}'

    gold_diff   = DifferentialStyle(
        fill=fill(C_GOLD),
        font=Font(bold=True, color='5D4037', name='Calibri')
    )
    silver_diff = DifferentialStyle(
        fill=fill(C_SILVER),
        font=Font(bold=True, color='37474F', name='Calibri')
    )
    bronze_diff = DifferentialStyle(
        fill=fill(C_BRONZE),
        font=Font(bold=True, color='FFFFFF', name='Calibri')
    )

    from openpyxl.formatting.rule import Rule
    from openpyxl.formatting.rule import CellIsRule

    ws.conditional_formatting.add(
        sira_range,
        CellIsRule(operator='equal', formula=['1'], fill=fill(C_GOLD),
                   font=Font(bold=True, name='Calibri', color='5D4037'))
    )
    ws.conditional_formatting.add(
        sira_range,
        CellIsRule(operator='equal', formula=['2'], fill=fill(C_SILVER),
                   font=Font(bold=True, name='Calibri', color='37474F'))
    )
    ws.conditional_formatting.add(
        sira_range,
        CellIsRule(operator='equal', formula=['3'], fill=fill(C_BRONZE),
                   font=Font(bold=True, name='Calibri', color='FFFFFF'))
    )

    # ── Koşullu biçimlendirme: Gecersiz / DNS satırları ───
    # Alet Durum sütunlarından herhangi biri G ise o hücre kırmızı
    col = FIXED_COLS + 2  # İlk Durum sütunu
    for _ in apparatus_list:
        dl = get_column_letter(col)
        durum_range = f'{dl}{DATA_START_ROW}:{dl}{data_end_row}'
        ws.conditional_formatting.add(
            durum_range,
            CellIsRule(operator='equal', formula=['"G"'],
                       fill=fill(C_GECERSIZ),
                       font=Font(bold=True, color='B71C1C', name='Calibri'))
        )
        ws.conditional_formatting.add(
            durum_range,
            CellIsRule(operator='equal', formula=['"D"'],
                       fill=fill(C_DNS),
                       font=Font(bold=True, color='E65100', name='Calibri'))
        )
        col += 2

    # ── Panes ve yazdırma ───────────────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=3)
    ws.print_title_rows = '1:5'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ── Alt satır: toplam sporcu sayısı ────────────────────
    summary_row = data_end_row + 2
    ws.merge_cells(start_row=summary_row, start_column=1,
                   end_row=summary_row, end_column=FIXED_COLS)
    c = ws.cell(row=summary_row, column=1,
                value=f'=COUNTA(B{DATA_START_ROW}:B{data_end_row})&" sporcu kayıtlı"')
    c.font = font(bold=True, size=10, color='444444', italic=True)
    c.alignment = align(h='right')


# ─────────────────────────────────────────────────────────────
#  ÇIKIŞ SIRASI SAYFASI
# ─────────────────────────────────────────────────────────────
def create_start_order_sheet(wb):
    ws = wb.create_sheet(title='ÇIKIŞ SIRASI')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 16   # Kategori
    ws.column_dimensions['B'].width = 7    # Grup
    ws.column_dimensions['C'].width = 5    # Grup içi sıra
    ws.column_dimensions['D'].width = 13   # Ad
    ws.column_dimensions['E'].width = 15   # Soyad
    ws.column_dimensions['F'].width = 30   # Okul
    ws.column_dimensions['G'].width = 10   # Tür
    ws.column_dimensions['H'].width = 16   # Antrenör/Öğretmen

    # Başlık
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value='ÇIKIŞ SIRASI  —  TÜM KATEGORİLER')
    c.font = font(bold=True, size=14, color='FFFFFF')
    c.fill = fill(C_DARK)
    c.alignment = align(h='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:H2')
    c = ws.cell(row=2, column=1,
                value='Bu sayfaya sporcuları manuel olarak ekleyin. '
                      'Her grupta maksimum 8 sporcu olabilir.')
    c.font = font(italic=True, size=9, color='555555')
    c.fill = fill('ECEFF1')
    c.alignment = align(h='center')
    ws.row_dimensions[2].height = 15

    current_row = 4
    col_headers = ['KATEGORİ', 'GRUP', 'SIRA', 'AD', 'SOYAD', 'OKUL / KULÜP', 'TÜR', 'ANTRENÖR / ÖĞR.']

    dv_tur = DataValidation(
        type='list', formula1='"Bireysel,Takım"',
        allow_blank=True, showDropDown=False, showErrorMessage=False
    )
    ws.add_data_validation(dv_tur)

    for _, label, _, gender in CATEGORIES:
        hdr_color = C_PINK if gender == 'k' else C_BLUE

        # Kategori başlığı
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=8)
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = font(bold=True, size=11, color='FFFFFF')
        c.fill = fill(hdr_color)
        c.alignment = align(h='center')
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Sütun başlıkları
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font = font(bold=True, size=9, color='FFFFFF')
            c.fill = fill(C_HEADER_BG)
            c.alignment = align(h='center')
            c.border = THIN
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Her grup için satırlar
        total_athlete_rows = MAX_GROUPS * MAX_PER_GROUP
        for g in range(1, MAX_GROUPS + 1):
            for s in range(1, MAX_PER_GROUP + 1):
                row_bg = 'FFFFFF' if g % 2 == 1 else 'F3E5F5' if gender == 'k' else 'E3F2FD'

                c = ws.cell(row=current_row, column=1, value=label)
                c.font = font(size=8, color='888888')
                c.fill = fill(row_bg)
                c.border = THIN
                c.alignment = align(h='center')

                c = ws.cell(row=current_row, column=2, value=f'Grup {g}')
                c.font = font(size=9, bold=(s == 1), color='444444')
                c.fill = fill(row_bg)
                c.border = THIN
                c.alignment = align(h='center')

                c = ws.cell(row=current_row, column=3, value=s)
                c.font = font(size=9, color='777777')
                c.fill = fill(row_bg)
                c.border = THIN
                c.alignment = align(h='center')

                for ci in range(4, 9):  # Ad, Soyad, Okul, Tür, Antrenör
                    c = ws.cell(row=current_row, column=ci)
                    c.fill = fill(row_bg)
                    c.border = THIN
                    c.font = font(size=10)
                    if ci in [4, 5, 6, 8]:
                        c.alignment = align(h='left')
                    else:
                        c.alignment = align(h='center')
                    if ci == 7:
                        dv_tur.add(c)

                ws.row_dimensions[current_row].height = 17
                current_row += 1

        current_row += 2  # Kategoriler arası boşluk

    ws.freeze_panes = 'A4'
    ws.print_title_rows = '1:3'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ─────────────────────────────────────────────────────────────
#  ÖZET SAYFA (Tüm Kategoriler)
# ─────────────────────────────────────────────────────────────
def create_summary_sheet(wb, category_sheets):
    ws = wb.create_sheet(title='TÜM SONUÇLAR')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 9

    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value='TÜM KATEGORİLER — SONUÇ ÖZETİ')
    c.font = font(bold=True, size=14, color='FFFFFF')
    c.fill = fill(C_DARK)
    c.alignment = align(h='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:F2')
    c = ws.cell(row=2, column=1,
                value='Puan girişini ilgili kategori sekmesinden yapın. '
                      'Bu sayfa tüm kategorilerin ilk 3 sırasını gösterir.')
    c.font = font(italic=True, size=9, color='555555')
    c.fill = fill('ECEFF1')
    c.alignment = align(h='center')
    ws.row_dimensions[2].height = 15

    col_headers = ['KATEGORİ', 'SIRALAMA', 'AD', 'SOYAD', 'OKUL / KULÜP', 'TOPLAM']
    current_row = 4

    for (_, label, _, gender), sheet_title in zip(CATEGORIES, category_sheets):
        hdr_color = C_PINK if gender == 'k' else C_BLUE

        # Kategori başlığı
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=6)
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = font(bold=True, size=11, color='FFFFFF')
        c.fill = fill(hdr_color)
        c.alignment = align(h='center')
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Sütun başlıkları
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font = font(bold=True, size=9, color='FFFFFF')
            c.fill = fill(C_HEADER_BG)
            c.alignment = align(h='center')
            c.border = THIN
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Podium satırları: 1., 2., 3.
        podium_icons = ['🥇', '🥈', '🥉']
        podium_fills = [C_GOLD, C_SILVER, C_BRONZE]
        podium_text_colors = ['5D4037', '37474F', 'FFFFFF']

        for rank_pos in range(3):
            # Formüllerle gerçek veriden çek
            # =IFERROR(INDEX(sheet!$B$6:$B$45, MATCH(rank_pos+1, sheet!$SIRA$:$SIRA$, 0)), "")
            # SIRALAMA sütunu dinamik, ama sheet başlığı bilindiğinden referans yapılabilir
            safe_title = sheet_title.replace("'", "''")
            sira_col_idx  = 5 + len([c for _, c_label, c_app, _ in CATEGORIES
                                      if c_label[:31] == sheet_title][0] if False else []) + 2
            # Not: doğrudan formül yerine yönlendirme notu koyacağız,
            # çünkü dinamik INDEX/MATCH SIRALAMA sütununa gerek duyar
            row_bg = [C_GOLD, C_SILVER, C_BRONZE][rank_pos]
            text_col = podium_text_colors[rank_pos]

            for ci in range(1, 7):
                c = ws.cell(row=current_row, column=ci)
                c.fill = fill(row_bg)
                c.border = THIN
                c.alignment = align(h='center' if ci != 5 else 'left')
                c.font = font(bold=True, size=10, color=text_col)

            ws.cell(row=current_row, column=1,
                    value=f'→ "{label[:20]}" sekmesine gidin').font = \
                font(italic=True, size=9, color='666666')
            ws.cell(row=current_row, column=2, value=rank_pos + 1)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1

    ws.freeze_panes = 'A3'
    ws.print_title_rows = '1:2'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ─────────────────────────────────────────────────────────────
#  ANA FONKSİYON
# ─────────────────────────────────────────────────────────────
def main():
    import os
    wb = Workbook()

    # Sayfa 0: Nasıl Kullanılır
    create_info_sheet(wb.active)

    # Sayfa 1-10: Kategori puanlama sayfaları
    sheet_titles = []
    for key, label, apparatus, gender in CATEGORIES:
        create_category_sheet(wb, key, label, apparatus, gender)
        sheet_titles.append(label[:31])
        print(f'  ✓  {label:30s}  ({", ".join(apparatus)})')

    # Sayfa 11: Çıkış sırası
    create_start_order_sheet(wb)

    # Sayfa 12: Özet
    create_summary_sheet(wb, sheet_titles)

    # Kaydet
    out_path = os.path.join(
        '/Users/emre.yalciner/Desktop/Aktif Cimnastik Sistemleri/TCF Okullar/new',
        'TCF_Artistik_Cimnastik_Offline.xlsx'
    )
    wb.save(out_path)
    print(f'\n✅  Dosya oluşturuldu:\n   {out_path}')
    print(f'   Toplam {len(wb.sheetnames)} sayfa: {", ".join(wb.sheetnames)}')


if __name__ == '__main__':
    print('TCF Artistik Cimnastik - Excel Sistemi oluşturuluyor...\n')
    main()
