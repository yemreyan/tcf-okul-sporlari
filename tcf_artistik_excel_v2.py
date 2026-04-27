#!/usr/bin/env python3
"""
TCF Artistik Cimnastik - Kapsamlı Offline Excel Sistemi v2
Gereksinim: pip install openpyxl
Çalıştırma: python3 tcf_artistik_excel_v2.py

Özellikler:
  • D + E − ND = Final hesaplama (Gecersiz/DNS sıfırlama)
  • Bireysel çok-alet sıralaması (eşit puan desteği)
  • Alet bazlı bireysel sıralama sütunları
  • Takım sıralaması (her alette en iyi 3 sporcu toplamı + takım kesintisi)
  • Alet Finalleri (ayrı tur, Top-8 formatı, kendi D/E/ND girişi)
  • Çıkış Sırası yönetimi
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

# ─────────────────────────────────────────────────────────────
#  KATEGORİ / ALET TANIMLAMALARI
# ─────────────────────────────────────────────────────────────
APPARATUS_LABELS = {
    'yer':      'Yer (FX)',          'atlama':    'Atlama (VT)',
    'barfiks':  'Barfiks (HB)',      'denge':     'Denge (BB)',
    'asimetrik':'As.Paralel (UB)',   'halka':     'Halka (SR)',
    'kulplu':   'Kulplu Beygir (PH)','paralel':   'Paralel (PB)',
    'serbest':  'Serbest (FX)',      'sirik':     'Sırık (PV)',
}
APP_SHORT = {
    'yer':'FX','atlama':'VT','barfiks':'HB','denge':'BB',
    'asimetrik':'UB','halka':'SR','kulplu':'PH','paralel':'PB',
    'serbest':'FX','sirik':'PV',
}

CATEGORIES = [
    ('minik_a_kiz',   'MİNİK A KIZLAR',   ['atlama','serbest'],                                           'k'),
    ('minik_a_erkek', 'MİNİK A ERKEKLER', ['yer','atlama'],                                               'e'),
    ('minik_b_kiz',   'MİNİK B KIZLAR',   ['atlama','serbest'],                                           'k'),
    ('minik_b_erkek', 'MİNİK B ERKEKLER', ['yer','atlama'],                                               'e'),
    ('kucuk_kiz',     'KÜÇÜK KIZLAR',     ['atlama','asimetrik','denge','serbest'],                       'k'),
    ('kucuk_erkek',   'KÜÇÜK ERKEKLER',   ['yer','atlama','paralel','barfiks'],                           'e'),
    ('yildiz_kiz',    'YILDIZ KIZLAR',    ['atlama','asimetrik','denge','serbest'],                       'k'),
    ('yildiz_erkek',  'YILDIZ ERKEKLER',  ['yer','kulplu','halka','atlama','paralel','barfiks'],          'e'),
    ('genc_kiz',      'GENÇ KIZLAR',      ['atlama','asimetrik','denge','serbest'],                       'k'),
    ('genc_erkek',    'GENÇ ERKEKLER',    ['yer','kulplu','halka','atlama','paralel','barfiks'],          'e'),
]

MAX_ATH   = 40    # Kategoride maksimum sporcu satırı
MAX_TEAMS = 20    # Takım bölümünde maksimum takım satırı
MAX_FIN   = 8     # Alet finalinde maksimum sporcu (Top-8)
DATA_ROW  = 6     # Sporcu verilerinin başlangıç satırı

# ─────────────────────────────────────────────────────────────
#  STİL SABİTLERİ
# ─────────────────────────────────────────────────────────────
C = dict(
    PINK='C2185B', BLUE='1565C0', DARK='1A237E', SUBHDR='263238',
    GOLD='FFD700', SILVER='B0BEC5', BRONZE='A1887F',
    FINAL='BBDEFB', TOPLAM='C8E6C9', RANK='FFF9C4',
    GECERSIZ='FFCDD2', DNS='FFF3E0', ALT='F9FBE7',
    TEAM_HDR='4A148C', TEAM_BG='F3E5F5',
    FIN_HDR='BF360C', FIN_BG='FFF8E1',
    NOTE='ECEFF1', WHITE='FFFFFF',
)

def _s(style='thin'):
    return Side(style=style)

def bdr(style='thin'):
    s = _s(style)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c):
    return PatternFill('solid', fgColor=c)

def fnt(bold=False, size=10, color='000000', italic=False):
    return Font(name='Calibri', bold=bold, size=size, color=color, italic=italic)

def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = bdr('thin')
MED  = bdr('medium')

def sc(ws, row, col, val=None, *, bold=False, sz=10, fg='000000',
       italic=False, bg=None, h='center', v='center', wrap=False,
       b=True, nf=None, merge_to=None):
    """Shortcut: style + set value to cell."""
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    c.font  = fnt(bold=bold, size=sz, color=fg, italic=italic)
    c.alignment = aln(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if b:
        c.border = THIN
    if nf:
        c.number_format = nf
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_to)
    return c

# ─────────────────────────────────────────────────────────────
#  SÜTUN YAPISI HESABI
# ─────────────────────────────────────────────────────────────
# Fixed cols: No(1) Ad(2) Soyad(3) Okul(4) Tür(5)
# Per apparatus: D(+1) E(+2) ND(+3) Final(+4) Dur(+5)
# Summary: TOPLAM, SIRALAMA
# Alet Sıra: one col per apparatus
FIXED = 5

def app_col(app_idx):
    """Returns dict with column indices for apparatus at app_idx (0-based)."""
    base = FIXED + app_idx * 5
    return {'D': base+1, 'E': base+2, 'ND': base+3, 'Final': base+4, 'Dur': base+5}

def summary_cols(n_app):
    base = FIXED + n_app * 5
    return {'Toplam': base+1, 'Sira': base+2, 'AppSiraStart': base+3}

def total_cols(n_app):
    return FIXED + n_app * 5 + 2 + n_app   # fixed + apparatus×5 + toplam+siralama + alet sıra cols

# ─────────────────────────────────────────────────────────────
#  1. REHBER SAYFASI
# ─────────────────────────────────────────────────────────────
def make_rehber(ws):
    ws.title = 'REHBER'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 68

    def row(r, icon, text, bold=False, bg=None, sz=10):
        ws.row_dimensions[r].height = 18
        b = ws.cell(row=r, column=2, value=icon)
        b.font = fnt(bold=True, size=sz)
        t = ws.cell(row=r, column=3, value=text)
        t.font = fnt(bold=bold, size=sz)
        t.alignment = aln(h='left')
        if bg:
            for col in [2, 3]:
                ws.cell(row=r, column=col).fill = fill(bg)

    ws.merge_cells('B1:C1')
    ws.row_dimensions[1].height = 38
    h = ws.cell(row=1, column=2,
                value='TCF ARTİSTİK CİMNASTİK  —  KAPSAMLI OFFLİNE PUAN SİSTEMİ  v2')
    h.font = fnt(bold=True, size=15, color='FFFFFF')
    h.fill = fill(C['DARK'])
    h.alignment = aln(h='center')

    ws.merge_cells('B2:C2')
    ws.row_dimensions[2].height = 20
    s = ws.cell(row=2, column=2,
                value='D + E − ND = Final  ·  Takım Sıralaması  ·  Alet Finalleri  ·  Çıkış Sırası')
    s.font = fnt(italic=True, size=10, color='444444')
    s.fill = fill('E8EAF6')
    s.alignment = aln(h='center')

    lines = [
        (3,  '',   ''),
        (4,  '📌', 'GENEL KULLANIM',                  True, 'E3F2FD', 12),
        (5,  '1.', 'YARIŞMA BİLGİLERİ sekmesine gidin → yarışma adı, tarih ve şehri doldurun.'),
        (6,  '2.', 'Kategori sekmelerinde (MİNİK A KIZLAR vb.) sporcu Ad/Soyad/Okul/Tür girin.'),
        (7,  '3.', 'Her alet için D, E, ND (Tarafsız Kesinti) puanlarını girin:'),
        (8,  '   ','    • Final = MAX(0 , D + E − ND)  →  otomatik hesaplanır (mavi sütun).'),
        (9,  '   ','    • Sadece Final puanı girecekseniz: D=0, E=final_puan, ND=0 kullanın.'),
        (10, '4.', 'Durum sütununa:  N = Normal  |  G = Gecersiz  |  D = Yarışmadı/DNS'),
        (11, '   ','    G veya D girilirse o alet Final puanı otomatik 0 yapılır.'),
        (12, '5.', 'TOPLAM ve SIRALAMA otomatik hesaplanır.  Eşit puanlılar aynı sırayı alır.'),
        (13, '6.', 'Her aletin kendi sıralama sütunu da vardır (alet bazlı final seçimi için).'),
        (14, '',   ''),
        (15, '🏆', 'TAKIM SIRALAMASI sekmesi',          True, 'EDE7F6', 12),
        (16, '•',  'Her kategoride takım adını A sütununa manuel girin.'),
        (17, '•',  'Her aletin "En İyi 3" değeri, ilgili kategori sayfasından otomatik çekilir.'),
        (18, '•',  'SUMPRODUCT formülü kullanılır.  Excel 2019+\'ta otomatik çalışır.'),
        (19, '•',  'Eski Excel (2016) kullanıyorsanız formül hücrelerini Ctrl+Shift+Enter ile onaylayın.'),
        (20, '•',  'Takım Kesintisi (ceza) sütununa manuel girin → Final takım puanı otomatik güncellenir.'),
        (21, '',   ''),
        (22, '🎯', 'ALET FİNALLERİ sekmesi',            True, 'FFF8E1', 12),
        (23, '•',  'Eleme sonucuna göre her alette en iyi 8 sporcu finals için el ile girilir.'),
        (24, '•',  'Final turu için ayrı D/E/ND puanları girilir → Final puanı yeniden hesaplanır.'),
        (25, '•',  'Final sıralaması yalnızca bu tur puanına göre otomatik hesaplanır.'),
        (26, '',   ''),
        (27, '⚠️', 'ÖNEMLİ NOTLAR',                    True, 'FFF8E1', 11),
        (28, '•',  'Ondalık için sistem ayarınıza göre nokta (.) veya virgül (,) kullanın.'),
        (29, '•',  'Düzenli kaydedin: Ctrl+S'),
        (30, '•',  '40\'dan fazla sporcu için yeni satır ekleyin; formülleri bir üst satırdan kopyalayın.'),
        (31, '',   ''),
        (32, '📋', 'KATEGORİLER VE ALETLER',            True, 'E8F5E9', 12),
        (33, '',   ''),
    ]
    for item in lines:
        r, icon, text = item[0], item[1], item[2]
        bold_ = item[3] if len(item) > 3 else False
        bg_   = item[4] if len(item) > 4 else None
        sz_   = item[5] if len(item) > 5 else 10
        row(r, icon, text, bold_, bg_, sz_)

    r = 34
    for _, lbl, apps, gender in CATEGORIES:
        icon = '👧' if gender == 'k' else '👦'
        bg_ = 'FCE4EC' if gender == 'k' else 'E3F2FD'
        app_str = ' + '.join(APPARATUS_LABELS.get(a, a) for a in apps)
        row(r, icon, f'{lbl}:  {app_str}', False, bg_)
        r += 1


# ─────────────────────────────────────────────────────────────
#  2. YARIŞMA BİLGİLERİ SAYFASI
# ─────────────────────────────────────────────────────────────
def make_yarisma(wb):
    ws = wb.create_sheet('YARIŞMA BİLGİLERİ')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45

    ws.merge_cells('B1:C1')
    ws.row_dimensions[1].height = 34
    h = ws.cell(row=1, column=2, value='YARIŞMA BİLGİLERİ')
    h.font = fnt(bold=True, size=14, color='FFFFFF')
    h.fill = fill(C['DARK'])
    h.alignment = aln(h='center')

    fields = [
        (3, 'Yarışma Adı',    'Buraya yarışma adını yazın'),
        (4, 'Tarih',           'GG.AA.YYYY'),
        (5, 'Şehir / İl',     ''),
        (6, 'Düzenleyen Kurum',''),
        (7, 'Hakem Başkanı',  ''),
    ]
    for r, label, placeholder in fields:
        ws.row_dimensions[r].height = 22
        lc = ws.cell(row=r, column=2, value=label)
        lc.font  = fnt(bold=True, size=11)
        lc.fill  = fill('E8EAF6')
        lc.alignment = aln(h='left')
        lc.border = THIN
        vc = ws.cell(row=r, column=3, value=placeholder)
        vc.font  = fnt(italic=True, size=11, color='666666')
        vc.alignment = aln(h='left')
        vc.border = THIN

    return ws


# ─────────────────────────────────────────────────────────────
#  3. KATEGORİ PUANLAMA SAYFASI
# ─────────────────────────────────────────────────────────────
def make_category(wb, key, label, apps, gender):
    ws = wb.create_sheet(title=label[:31])
    ws.sheet_view.showGridLines = False

    n_app    = len(apps)
    s_cols   = summary_cols(n_app)
    t_cols   = total_cols(n_app)
    hdr_bg   = C['PINK'] if gender == 'k' else C['BLUE']
    end_data = DATA_ROW + MAX_ATH - 1

    # ── Sütun genişlikleri ──────────────────────────────────
    ws.column_dimensions['A'].width = 4.5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 10
    for i in range(n_app):
        ac = app_col(i)
        for ci, w in zip([ac['D'], ac['E'], ac['ND'], ac['Final'], ac['Dur']],
                         [6.5, 6.5, 6.5, 9, 6]):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions[get_column_letter(s_cols['Toplam'])].width = 11
    ws.column_dimensions[get_column_letter(s_cols['Sira'])].width   = 9
    for i in range(n_app):
        ws.column_dimensions[get_column_letter(s_cols['AppSiraStart']+i)].width = 7

    # ── Satır 1: Başlık ─────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=t_cols)
    ws.row_dimensions[1].height = 32
    h1 = ws.cell(row=1, column=1, value=label)
    h1.font = fnt(bold=True, size=15, color='FFFFFF')
    h1.fill = fill(hdr_bg)
    h1.alignment = aln(h='center')

    # ── Satır 2: Yarışma bilgisi referansı ──────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=t_cols)
    ws.row_dimensions[2].height = 15
    c2 = ws.cell(row=2, column=1,
                 value="='YARIŞMA BİLGİLERİ'!C3&\"  —  \"&'YARIŞMA BİLGİLERİ'!C4&\"  |  \"&'YARIŞMA BİLGİLERİ'!C5")
    c2.font = fnt(italic=True, size=9, color='555555')
    c2.fill = fill(C['NOTE'])
    c2.alignment = aln(h='center')

    # ── Satır 3: Alet grup başlıkları (merge) ───────────────
    ws.row_dimensions[3].height = 22
    for ci in range(1, FIXED+1):
        c = ws.cell(row=3, column=ci)
        c.fill = fill(hdr_bg)
        c.border = THIN
    for i, app in enumerate(apps):
        ac = app_col(i)
        ws.merge_cells(start_row=3, start_column=ac['D'], end_row=3, end_column=ac['Dur'])
        c = ws.cell(row=3, column=ac['D'],
                    value=APPARATUS_LABELS.get(app, app).upper())
        c.font = fnt(bold=True, size=10, color='FFFFFF')
        c.fill = fill(hdr_bg)
        c.alignment = aln(h='center')
        c.border = THIN
        for ci in range(ac['D']+1, ac['Dur']+1):
            ws.cell(row=3, column=ci).border = THIN
            ws.cell(row=3, column=ci).fill = fill(hdr_bg)
    for ci in [s_cols['Toplam'], s_cols['Sira']]:
        c = ws.cell(row=3, column=ci)
        c.fill = fill(hdr_bg)
        c.border = THIN
    ws.merge_cells(start_row=3, start_column=s_cols['AppSiraStart'],
                   end_row=3, end_column=s_cols['AppSiraStart']+n_app-1)
    c = ws.cell(row=3, column=s_cols['AppSiraStart'],
                value='ALET SIRALARI  (Eleme Seçimi için)')
    c.font  = fnt(bold=True, size=9, color='FFFFFF')
    c.fill  = fill(hdr_bg)
    c.alignment = aln(h='center')
    for ci in range(s_cols['AppSiraStart']+1, s_cols['AppSiraStart']+n_app):
        ws.cell(row=3, column=ci).fill  = fill(hdr_bg)
        ws.cell(row=3, column=ci).border = THIN

    # ── Satır 4: Sütun alt başlıkları ───────────────────────
    ws.row_dimensions[4].height = 30
    sub_hdrs = ['No','Ad','Soyad','Okul / Kulüp','Tür']
    for i, app in enumerate(apps):
        sub_hdrs += ['D','E','ND','Final','Dur']
    sub_hdrs += ['TOPLAM','SIRALAMA']
    sub_hdrs += [APP_SHORT.get(a, a[:3]) for a in apps]

    for ci, h in enumerate(sub_hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = fnt(bold=True, size=9, color='FFFFFF')
        c.fill = fill(C['SUBHDR'])
        c.alignment = aln(h='center', wrap=True)
        c.border = THIN

    # ── Satır 5: Not ────────────────────────────────────────
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=t_cols)
    ws.row_dimensions[5].height = 13
    n5 = ws.cell(row=5, column=1,
                 value='Final = MAX(0, D+E−ND)  |  Dur: N=Normal · G=Gecersiz · D=Yarışmadı/DNS  '
                       '|  Mavi sütun = formül (elle girme)  |  Yeşil = Toplam  |  Sarı = Sıralama')
    n5.font = fnt(italic=True, size=8, color='444444')
    n5.fill = fill(C['NOTE'])
    n5.alignment = aln(h='center')

    # ── Data Validation: Tür dropdown ───────────────────────
    dv_tur = DataValidation(type='list', formula1='"Bireysel,Takım"',
                             allow_blank=True, showDropDown=False, showErrorMessage=False)
    ws.add_data_validation(dv_tur)
    dv_dur = DataValidation(type='list', formula1='"N,G,D"',
                             allow_blank=True, showDropDown=False, showErrorMessage=False)
    ws.add_data_validation(dv_dur)

    # ── Veri Satırları ───────────────────────────────────────
    toplam_range = (f'${get_column_letter(s_cols["Toplam"])}${DATA_ROW}:'
                    f'${get_column_letter(s_cols["Toplam"])}${end_data}')

    for row_i, row in enumerate(range(DATA_ROW, end_data+1)):
        bg_row = C['WHITE'] if row_i % 2 == 0 else C['ALT']

        # No, Ad, Soyad, Okul
        sc(ws, row, 1, row_i+1, sz=9, fg='888888', bg=bg_row, h='center')
        for ci in [2, 3]:
            sc(ws, row, ci, bg=bg_row, h='left')
        sc(ws, row, 4, bg=bg_row, h='left', sz=9)

        # Tür
        c = sc(ws, row, 5, bg=bg_row, h='center')
        dv_tur.add(c)

        # Alet sütunları
        final_refs = []
        for i, app in enumerate(apps):
            ac = app_col(i)
            dL  = get_column_letter(ac['D'])
            eL  = get_column_letter(ac['E'])
            ndL = get_column_letter(ac['ND'])
            fL  = get_column_letter(ac['Final'])
            durL= get_column_letter(ac['Dur'])

            # D, E, ND (giriş)
            for ci in [ac['D'], ac['E'], ac['ND']]:
                sc(ws, row, ci, bg=bg_row, h='center', nf='0.000')

            # Final (formül, mavi bg)
            final_formula = (
                f'=IF(OR({durL}{row}="G",{durL}{row}="D"),0,'
                f'IF(AND({dL}{row}="",{eL}{row}=""),0,'
                f'MAX(0,IFERROR({dL}{row}+0,0)+IFERROR({eL}{row}+0,0)'
                f'-IFERROR({ndL}{row}+0,0))))'
            )
            sc(ws, row, ac['Final'], final_formula,
               bold=True, bg=C['FINAL'], h='center', nf='0.000')

            # Durum
            c = sc(ws, row, ac['Dur'], bg=bg_row, h='center', sz=9)
            dv_dur.add(c)
            final_refs.append((fL, durL))

        # TOPLAM
        eff = '+'.join(
            f'IF(OR({dL}{row}="G",{dL}{row}="D"),0,{fL}{row})'
            for fL, dL in final_refs
        )
        toplam_formula = f'=IF(B{row}="","",{eff})'
        sc(ws, row, s_cols['Toplam'], toplam_formula,
           bold=True, bg=C['TOPLAM'], h='center', nf='0.000')

        # SIRALAMA
        siralama_formula = (
            f'=IF(OR(B{row}="",{get_column_letter(s_cols["Toplam"])}{row}=0),'
            f'"-",RANK.EQ({get_column_letter(s_cols["Toplam"])}{row},{toplam_range},0))'
        )
        sc(ws, row, s_cols['Sira'], siralama_formula,
           bold=True, bg=C['RANK'], h='center')

        # Alet bazlı sıralama sütunları
        for i, app in enumerate(apps):
            ac   = app_col(i)
            fL   = get_column_letter(ac['Final'])
            durL = get_column_letter(ac['Dur'])
            app_range = f'${fL}${DATA_ROW}:${fL}${end_data}'
            app_sira_formula = (
                f'=IF(OR(B{row}="",{fL}{row}=0),"-",'
                f'IF(OR({durL}{row}="G",{durL}{row}="D"),'
                f'IF({durL}{row}="G","G","D"),'
                f'COUNTIF({app_range},">"&{fL}{row})+1))'
            )
            sc(ws, row, s_cols['AppSiraStart']+i, app_sira_formula,
               sz=9, bg=bg_row, h='center')

    # ── Koşullu biçimlendirme ────────────────────────────────
    sira_letter = get_column_letter(s_cols['Sira'])
    sira_range  = f'{sira_letter}{DATA_ROW}:{sira_letter}{end_data}'
    ws.conditional_formatting.add(sira_range, CellIsRule(
        operator='equal', formula=['1'],
        fill=fill(C['GOLD']), font=Font(bold=True, name='Calibri', color='5D4037')))
    ws.conditional_formatting.add(sira_range, CellIsRule(
        operator='equal', formula=['2'],
        fill=fill(C['SILVER']), font=Font(bold=True, name='Calibri', color='37474F')))
    ws.conditional_formatting.add(sira_range, CellIsRule(
        operator='equal', formula=['3'],
        fill=fill(C['BRONZE']), font=Font(bold=True, name='Calibri', color='FFFFFF')))

    # Durum sütunları renklendirme
    for i in range(n_app):
        ac    = app_col(i)
        durL  = get_column_letter(ac['Dur'])
        d_rng = f'{durL}{DATA_ROW}:{durL}{end_data}'
        ws.conditional_formatting.add(d_rng, CellIsRule(
            operator='equal', formula=['"G"'],
            fill=fill(C['GECERSIZ']), font=Font(bold=True, name='Calibri', color='B71C1C')))
        ws.conditional_formatting.add(d_rng, CellIsRule(
            operator='equal', formula=['"D"'],
            fill=fill(C['DNS']), font=Font(bold=True, name='Calibri', color='E65100')))

    # ── Dondur & Yazdır ─────────────────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_ROW, column=3)
    ws.print_title_rows = '1:5'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    for m in ['left','right','top','bottom']:
        setattr(ws.page_margins, m, 0.4 if m in ['left','right'] else 0.5)

    # ── Sporcu sayısı ────────────────────────────────────────
    stat_row = end_data + 2
    ws.merge_cells(start_row=stat_row, start_column=1,
                   end_row=stat_row, end_column=FIXED)
    c = ws.cell(row=stat_row, column=1,
                value=f'=COUNTA(B{DATA_ROW}:B{end_data})&" sporcu kayıtlı"')
    c.font = fnt(italic=True, size=9, color='777777', bold=True)
    c.alignment = aln(h='right')

    # ── Takım Sıralaması bölümü (ayrı sheet'te ele alınıyor) ─
    # (TAKIM sheet'te cross-sheet formüller kullanılıyor)

    return ws, n_app, label, apps


# ─────────────────────────────────────────────────────────────
#  4. TAKIM SIRALAMASI SAYFASI
# ─────────────────────────────────────────────────────────────
def make_takim(wb, category_info):
    """
    category_info: list of (label, apps, gender) for each category
    Formüller ilgili kategori sayfasından SUMPRODUCT ile çeker.
    """
    ws = wb.create_sheet('TAKIM SIRALAMASI')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 34
    h = ws.cell(row=1, column=1, value='TAKIM SIRALAMASI  —  TÜM KATEGORİLER')
    h.font = fnt(bold=True, size=14, color='FFFFFF')
    h.fill = fill(C['DARK'])
    h.alignment = aln(h='center')

    ws.merge_cells('A2:H2')
    ws.row_dimensions[2].height = 16
    n = ws.cell(row=2, column=1,
                value='Takım adını A sütununa girin. En İyi 3 değerleri otomatik çekilir. '
                      'Eski Excel (2016) için formül hücrelerini Ctrl+Shift+Enter ile onaylayın.')
    n.font = fnt(italic=True, size=9, color='555555')
    n.fill = fill(C['NOTE'])
    n.alignment = aln(h='center')

    current_row = 4

    for label, apps, gender in category_info:
        n_app    = len(apps)
        hdr_bg   = C['PINK'] if gender == 'k' else C['BLUE']
        max_cols = 1 + n_app + 3  # Takım + apps + Toplam + Kesinti + Final + Sıra

        # Dinamik sütun genişlikleri
        ws.column_dimensions['A'].width = 28
        col_letters = [get_column_letter(c) for c in range(2, 2+n_app+4)]
        for cl in col_letters:
            ws.column_dimensions[cl].width = 11

        # Kategori başlığı
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=max_cols)
        ws.row_dimensions[current_row].height = 24
        ch = ws.cell(row=current_row, column=1, value=label + '  —  TAKIM SIRALAMASI')
        ch.font = fnt(bold=True, size=11, color='FFFFFF')
        ch.fill = fill(hdr_bg)
        ch.alignment = aln(h='center')
        current_row += 1

        # Sütun başlıkları
        ws.row_dimensions[current_row].height = 22
        hdrs = ['TAKIM ADI'] + \
               [f'En İyi 3\n{APP_SHORT.get(a,a[:4])}' for a in apps] + \
               ['TOPLAM', 'KESİNTİ', 'FİNAL PUAN', 'SIRALAMA']
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font = fnt(bold=True, size=9, color='FFFFFF')
            c.fill = fill(C['SUBHDR'])
            c.alignment = aln(h='center', wrap=True)
            c.border = THIN
        current_row += 1

        team_data_start = current_row
        team_data_end   = current_row + MAX_TEAMS - 1
        toplam_col_idx  = 1 + n_app + 1   # Takım(1) + apps(n_app) + Toplam
        kesinti_col_idx = toplam_col_idx + 1
        final_col_idx   = kesinti_col_idx + 1
        sira_col_idx    = final_col_idx + 1

        toplam_letter = get_column_letter(toplam_col_idx)
        final_letter  = get_column_letter(final_col_idx)
        sira_letter   = get_column_letter(sira_col_idx)
        final_range   = (f'${final_letter}${team_data_start}:'
                         f'${final_letter}${team_data_end}')

        for ti, t_row in enumerate(range(team_data_start, team_data_end+1)):
            bg_ = C['WHITE'] if ti % 2 == 0 else C['TEAM_BG']

            # Takım adı (manuel giriş)
            c = ws.cell(row=t_row, column=1)
            c.font      = fnt(size=10, bold=True)
            c.fill      = fill(bg_)
            c.alignment = aln(h='left')
            c.border    = THIN

            # En İyi 3 formülleri per alet
            app_sum_cols = []
            for ai, app in enumerate(apps):
                ac    = app_col(ai)
                f_col = get_column_letter(ac['Final'])
                d_col = 'D'   # Okul/Kulüp sütunu sabit D
                e_col = 'E'   # Tür sütunu sabit E
                data_start = DATA_ROW
                data_end   = DATA_ROW + MAX_ATH - 1
                sheet_ref  = label[:31].replace("'", "''")

                formula = (
                    f"=IFERROR(SUMPRODUCT(LARGE(IF(('{sheet_ref}'!$D${data_start}:$D${data_end}=A{t_row})"
                    f"*('{sheet_ref}'!$E${data_start}:$E${data_end}=\"Takım\"),"
                    f"'{sheet_ref}'!${f_col}${data_start}:${f_col}${data_end},0),{{1,2,3}})),0)"
                )
                ci = 1 + ai + 1
                c = ws.cell(row=t_row, column=ci, value=formula)
                c.font      = fnt(size=10)
                c.fill      = fill(bg_)
                c.alignment = aln(h='center')
                c.border    = THIN
                c.number_format = '0.000'
                app_sum_cols.append(get_column_letter(ci))

            # TOPLAM
            toplam_formula = '=' + '+'.join(f'{lc}{t_row}' for lc in app_sum_cols)
            c = ws.cell(row=t_row, column=toplam_col_idx, value=toplam_formula)
            c.font = fnt(bold=True, size=10)
            c.fill = fill(C['TOPLAM'])
            c.alignment = aln(h='center')
            c.border = THIN
            c.number_format = '0.000'

            # Kesinti (manuel)
            c = ws.cell(row=t_row, column=kesinti_col_idx, value=0)
            c.font = fnt(size=10)
            c.fill = fill(C['GECERSIZ'])
            c.alignment = aln(h='center')
            c.border = THIN
            c.number_format = '0.000'

            # Final Puan
            final_formula = (
                f'=IF(A{t_row}="","",{toplam_letter}{t_row}'
                f'-IFERROR({get_column_letter(kesinti_col_idx)}{t_row}+0,0))'
            )
            c = ws.cell(row=t_row, column=final_col_idx, value=final_formula)
            c.font = fnt(bold=True, size=10)
            c.fill = fill(C['TOPLAM'])
            c.alignment = aln(h='center')
            c.border = THIN
            c.number_format = '0.000'

            # Siralama
            sira_formula = (
                f'=IF(A{t_row}="","-",'
                f'RANK.EQ({final_letter}{t_row},{final_range},0))'
            )
            c = ws.cell(row=t_row, column=sira_col_idx, value=sira_formula)
            c.font = fnt(bold=True, size=10)
            c.fill = fill(C['RANK'])
            c.alignment = aln(h='center')
            c.border = THIN

        # Altın/gümüş/bronz takım sırası
        sira_rng = f'{sira_letter}{team_data_start}:{sira_letter}{team_data_end}'
        ws.conditional_formatting.add(sira_rng, CellIsRule(
            operator='equal', formula=['1'],
            fill=fill(C['GOLD']), font=Font(bold=True, name='Calibri', color='5D4037')))
        ws.conditional_formatting.add(sira_rng, CellIsRule(
            operator='equal', formula=['2'],
            fill=fill(C['SILVER']), font=Font(bold=True, name='Calibri', color='37474F')))
        ws.conditional_formatting.add(sira_rng, CellIsRule(
            operator='equal', formula=['3'],
            fill=fill(C['BRONZE']), font=Font(bold=True, name='Calibri', color='FFFFFF')))

        current_row = team_data_end + 3

    ws.freeze_panes = 'A3'
    ws.print_title_rows = '1:2'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


# ─────────────────────────────────────────────────────────────
#  5. ALET FİNALLERİ SAYFASI
# ─────────────────────────────────────────────────────────────
def make_alet_finalleri(wb):
    ws = wb.create_sheet('ALET FİNALLERİ')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 6    # Eleme Sırası
    ws.column_dimensions['B'].width = 12   # Ad
    ws.column_dimensions['C'].width = 14   # Soyad
    ws.column_dimensions['D'].width = 26   # Okul
    ws.column_dimensions['E'].width = 7    # D
    ws.column_dimensions['F'].width = 7    # E
    ws.column_dimensions['G'].width = 7    # ND
    ws.column_dimensions['H'].width = 9    # Final (formül)
    ws.column_dimensions['I'].width = 8    # Durum
    ws.column_dimensions['J'].width = 9    # Final Sıralama

    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 34
    h = ws.cell(row=1, column=1, value='ALET FİNALLERİ  —  AYRICALIKLI TUR PUANLAMASI  (Top-8)')
    h.font = fnt(bold=True, size=14, color='FFFFFF')
    h.fill = fill(C['DARK'])
    h.alignment = aln(h='center')

    ws.merge_cells('A2:J2')
    ws.row_dimensions[2].height = 16
    n = ws.cell(row=2, column=1,
                value='Eleme sırasına göre her alette top-8 sporcu elle girilir. '
                      'Final turu için D/E/ND puanları girilir → Final ve Sıralama otomatik hesaplanır.')
    n.font = fnt(italic=True, size=9, color='444444')
    n.fill = fill(C['NOTE'])
    n.alignment = aln(h='center')

    current_row = 4
    col_hdrs = ['Eleme\nSırası', 'Ad', 'Soyad', 'Okul / Kulüp', 'D', 'E', 'ND', 'FİNAL', 'Dur', 'FİN.SIRA']

    dv_dur = DataValidation(type='list', formula1='"N,G,D"',
                             allow_blank=True, showDropDown=False, showErrorMessage=False)
    ws.add_data_validation(dv_dur)

    for _, cat_label, apps, gender in CATEGORIES:
        hdr_bg = C['PINK'] if gender == 'k' else C['BLUE']

        for app in apps:
            # Bölüm başlığı
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=10)
            ws.row_dimensions[current_row].height = 26
            c = ws.cell(row=current_row, column=1,
                        value=f'{APPARATUS_LABELS.get(app,app).upper()}  —  {cat_label}')
            c.font = fnt(bold=True, size=11, color='FFFFFF')
            c.fill = fill(C['FIN_HDR'])
            c.alignment = aln(h='center')
            current_row += 1

            # Sütun başlıkları
            ws.row_dimensions[current_row].height = 26
            for ci, h in enumerate(col_hdrs, 1):
                c = ws.cell(row=current_row, column=ci, value=h)
                c.font = fnt(bold=True, size=9, color='FFFFFF')
                c.fill = fill(C['SUBHDR'])
                c.alignment = aln(h='center', wrap=True)
                c.border = THIN
            current_row += 1

            fin_start = current_row
            fin_end   = current_row + MAX_FIN - 1
            fin_range = f'$H${fin_start}:$H${fin_end}'

            for fi in range(MAX_FIN):
                row = current_row + fi
                bg_ = C['WHITE'] if fi % 2 == 0 else C['FIN_BG']
                ws.row_dimensions[row].height = 18

                sc(ws, row, 1, fi+1, sz=9, fg='888888', bg=bg_, h='center')   # Eleme Sırası
                for ci in [2, 3, 4]:
                    sc(ws, row, ci, bg=bg_, h='left' if ci > 1 else 'center', sz=10 if ci < 4 else 9)
                for ci in [5, 6, 7]:  # D, E, ND
                    sc(ws, row, ci, bg=bg_, h='center', nf='0.000')

                # Final (formül)
                final_f = (
                    f'=IF(OR(I{row}="G",I{row}="D"),0,'
                    f'IF(AND(E{row}="",F{row}=""),0,'
                    f'MAX(0,IFERROR(E{row}+0,0)+IFERROR(F{row}+0,0)-IFERROR(G{row}+0,0))))'
                )
                sc(ws, row, 8, final_f, bold=True, bg=C['FINAL'], h='center', nf='0.000')

                # Durum
                c = sc(ws, row, 9, bg=bg_, h='center', sz=9)
                dv_dur.add(c)

                # Final Sıralama
                sira_f = (
                    f'=IF(OR(B{row}="",H{row}=0),"-",'
                    f'RANK.EQ(H{row},{fin_range},0))'
                )
                sc(ws, row, 10, sira_f, bold=True, bg=C['RANK'], h='center')

            # Koşullu biçimlendirme
            ws.conditional_formatting.add(f'J{fin_start}:J{fin_end}', CellIsRule(
                operator='equal', formula=['1'],
                fill=fill(C['GOLD']), font=Font(bold=True, name='Calibri', color='5D4037')))
            ws.conditional_formatting.add(f'J{fin_start}:J{fin_end}', CellIsRule(
                operator='equal', formula=['2'],
                fill=fill(C['SILVER']), font=Font(bold=True, name='Calibri', color='37474F')))
            ws.conditional_formatting.add(f'J{fin_start}:J{fin_end}', CellIsRule(
                operator='equal', formula=['3'],
                fill=fill(C['BRONZE']), font=Font(bold=True, name='Calibri', color='FFFFFF')))

            current_row = fin_end + 3

    ws.freeze_panes = 'B3'
    ws.print_title_rows = '1:2'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


# ─────────────────────────────────────────────────────────────
#  6. ÇIKIŞ SIRASI SAYFASI
# ─────────────────────────────────────────────────────────────
def make_cikis(wb):
    ws = wb.create_sheet('ÇIKIŞ SIRASI')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 16

    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 32
    h = ws.cell(row=1, column=1, value='ÇIKIŞ SIRASI  —  TÜM KATEGORİLER')
    h.font = fnt(bold=True, size=14, color='FFFFFF')
    h.fill = fill(C['DARK'])
    h.alignment = aln(h='center')

    ws.merge_cells('A2:H2')
    ws.row_dimensions[2].height = 14
    n = ws.cell(row=2, column=1,
                value='Sporcuları manuel olarak gruplara atayın. Maksimum 8 sporcu / grup.')
    n.font = fnt(italic=True, size=9, color='555555')
    n.fill = fill(C['NOTE'])
    n.alignment = aln(h='center')

    current_row = 4
    col_hdrs = ['KATEGORİ', 'GRUP', 'SIRA', 'AD', 'SOYAD', 'OKUL / KULÜP', 'TÜR', 'ANTRENÖR / ÖĞR.']
    hdr_bg = C['SUBHDR']

    dv_tur = DataValidation(type='list', formula1='"Bireysel,Takım"',
                             allow_blank=True, showDropDown=False, showErrorMessage=False)
    ws.add_data_validation(dv_tur)

    for _, label, _, gender in CATEGORIES:
        cat_bg = C['PINK'] if gender == 'k' else C['BLUE']

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=8)
        ws.row_dimensions[current_row].height = 22
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = fnt(bold=True, size=11, color='FFFFFF')
        c.fill = fill(cat_bg)
        c.alignment = aln(h='center')
        current_row += 1

        for ci, h in enumerate(col_hdrs, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font = fnt(bold=True, size=9, color='FFFFFF')
            c.fill = fill(hdr_bg)
            c.alignment = aln(h='center')
            c.border = THIN
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for g in range(1, 9):   # 8 grup
            for s in range(1, 9):  # 8 sporcu per grup
                bg_ = C['WHITE'] if g % 2 == 1 else ('F3E5F5' if gender == 'k' else 'E3F2FD')
                ws.row_dimensions[current_row].height = 17

                c = ws.cell(row=current_row, column=1, value=label)
                c.font = fnt(size=8, color='888888'); c.fill = fill(bg_); c.border = THIN; c.alignment = aln(h='center')
                c = ws.cell(row=current_row, column=2, value=f'Grup {g}')
                c.font = fnt(size=9, bold=(s==1), color='444444'); c.fill = fill(bg_); c.border = THIN; c.alignment = aln(h='center')
                c = ws.cell(row=current_row, column=3, value=s)
                c.font = fnt(size=9, color='777777'); c.fill = fill(bg_); c.border = THIN; c.alignment = aln(h='center')

                for ci in range(4, 9):
                    c = ws.cell(row=current_row, column=ci)
                    c.fill = fill(bg_); c.border = THIN
                    c.font = fnt(size=10)
                    c.alignment = aln(h='left' if ci != 7 else 'center')
                    if ci == 7:
                        dv_tur.add(c)

                current_row += 1
        current_row += 2

    ws.freeze_panes = 'A4'
    ws.print_title_rows = '1:3'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


# ─────────────────────────────────────────────────────────────
#  ANA PROGRAM
# ─────────────────────────────────────────────────────────────
def main():
    wb = Workbook()

    print('TCF Artistik Cimnastik - Kapsamlı Excel Sistemi v2 oluşturuluyor...\n')

    # Rehber (aktif ilk sayfa)
    make_rehber(wb.active)
    print('  ✓  REHBER')

    # Yarışma bilgileri
    make_yarisma(wb)
    print('  ✓  YARIŞMA BİLGİLERİ')

    # Kategori puanlama sayfaları
    category_info = []
    for key, label, apps, gender in CATEGORIES:
        make_category(wb, key, label, apps, gender)
        category_info.append((label, apps, gender))
        print(f'  ✓  {label:30s}  ({", ".join(APPARATUS_LABELS.get(a,a) for a in apps)})')

    # Takım sıralaması
    make_takim(wb, category_info)
    print('  ✓  TAKIM SIRALAMASI')

    # Alet finalleri
    make_alet_finalleri(wb)
    print('  ✓  ALET FİNALLERİ')

    # Çıkış sırası
    make_cikis(wb)
    print('  ✓  ÇIKIŞ SIRASI')

    out = os.path.join(
        '/Users/emre.yalciner/Desktop/Aktif Cimnastik Sistemleri/TCF Okullar/new',
        'TCF_Artistik_Cimnastik_v2.xlsx'
    )
    wb.save(out)
    print(f'\n✅  Dosya oluşturuldu:\n   {out}')
    print(f'   Toplam {len(wb.sheetnames)} sayfa: {", ".join(wb.sheetnames)}')


if __name__ == '__main__':
    main()
